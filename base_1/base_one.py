import openpyxl
import pprint as ppr

# この実行ファイルのパス
cur_path = "/Users/kashimataichi/play_python/excel_python/base_1/"
excel_path = cur_path + "Sample.xlsx"

# エクセルファイルの読み込み
wb = openpyxl.load_workbook(excel_path)

# ブックのシート名の一覧を出力する
print(wb.sheetnames)
print(type(wb.sheetnames))

# ブックの1枚目のシートの名称を出力
print(wb.worksheets[0])
print(type(wb.worksheets[0]))

# 値の書き込み
sheet = wb['Sheet1']
sheet['A1'] = 'QQQ'
# 実施後に上書き保存をしないと結果が反映されないので注意
wb.save(excel_path)

# セルオブジェクトの取得
cell = sheet['A1']
print(type(cell))
print(cell.value)

# 任意の範囲を2次元タプルで取得
sheet2 = wb['Sheet2']
ppr.pprint(sheet2['A1:C3'])

ppr.pprint(sheet2['A1:C3'][0])
ppr.pprint(sheet2['A1:C3'][1])
ppr.pprint(sheet2['A1:C3'][2])

ppr.pprint(sheet2['A1:C3'][0][0].value)
ppr.pprint(sheet2['A1:C3'][0][1].value)
ppr.pprint(sheet2['A1:C3'][0][2].value)
