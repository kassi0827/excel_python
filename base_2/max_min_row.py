import openpyxl

# この実行ファイルのパス
cur_path = "/Users/kashimataichi/play_python/excel_python/base_2/"
excel_path = cur_path + "Book2.xlsx"

wb = openpyxl.load_workbook(excel_path)

ws = wb['max_min_row']

# データの最終行と最終列
print(ws.max_row)
print(ws.max_column)
